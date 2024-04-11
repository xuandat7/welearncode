from flask import Flask,redirect,url_for,render_template,request,session,flash,jsonify,session
# import pyodbc
import random
import smtplib
from email.mime.text import MIMEText
import pandas as pd
from fuzzywuzzy import process  # type: ignore
from pyvi import ViTokenizer # type: ignore
import re
from sklearn.feature_extraction.text import CountVectorizer
import pickle
import mysql.connector
from openpyxl import load_workbook
from datetime import timedelta

app=Flask(__name__)
#Config session
app.config["SECRET_KEY"]="quanhoangduong"
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=5)
app.config['SESSION_COOKIE_NAME'] = 'Group3WeLearnCodeProject'
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True

conn=mysql.connector.connect(user='root',password='123456',host='localhost')
cursor = conn.cursor()

questions = []
# cursor.execute("SELECT DISTINCT tenmon FROM btlweb.nganhangcauhoi")
# res=cursor.fetchall()
# MON_HOC=[str(row[0]) for row in res]


#khoi tao trang web dau tien 
@app.route('/')
def index():
    session.pop("user", None)
    return render_template('trangkhoidau.html')

@app.route('/home')
def home():
    if "user" in session:
        return render_template("trangkhoidau_logged_in.html")
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
 
@app.route('/admin')
def admin():
    if "user" in session:
        return render_template("admin.html")  
    else: 
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
    
#Đăng nhập và xử lý đăng nhâp
def check(a,b):
    if a == "admin@123.com" and b == "123456":
        return 2   
    cursor.execute("SELECT * FROM btlweb.nguoidung")
    result = cursor.fetchall()
    if result is not None:
        for row in result:
            if row[1] == a and row[2] == b:
                return 1
    return 0




@app.route('/login', methods=["POST", "GET"] )
def login():
    if request.method=="POST":
        tendangnhap=request.form["username"]
        matkhau=request.form["password"]
        if check(tendangnhap,matkhau) == 1:
            session["user"]=check(tendangnhap,matkhau)
            return redirect(url_for("home"))
        elif check(tendangnhap, matkhau) == 2:
            session["user"]=check(tendangnhap,matkhau)
            return redirect(url_for("admin"))
        else:
            flash("Tài khoản hoặc mật khẩu của bạn bị sai",category="info")
            return render_template("login.html")
    else:
        return render_template("login.html")


# phần xử lý cho việc người dùng nhập tên môn để làm bai trắc nghiệm

# def search_similar_names(name):
#     conn=mysql.connector.connect(user='root',password='123456',host='localhost')
#     cursor = conn.cursor()
#     # Lấy tất cả các tên từ cơ sở dữ liệu
#     cursor.execute("SELECT DISTINCT tenmon FROM btlweb.nganhangcauhoi")
#     res=cursor.fetchall()
#     all_names=[str(row[0]) for row in res]
#     # Sử dụng fuzzy matching để tìm các tên gần giống
#     similar_names = process.extract(name, all_names, limit=1)
#     conn.close()
#     return similar_names

def search_name_in_database(name):
    # Thực hiện truy vấn SQL để tìm kiếm tên trong cơ sở dữ liệu
    cursor.execute("SELECT * FROM btlweb.nganhangcauhoi WHERE tenmon=%s", (name,))
    result = cursor.fetchall()
    return result

def search_subjects(keyword):
    cursor.execute("SELECT DISTINCT tenmon FROM btlweb.nganhangcauhoi")
    res=cursor.fetchall()
    subjects=[str(row[0]) for row in res]
    matched_subjects = [subject for subject in subjects if keyword.lower() in subject.lower()]
    return matched_subjects

@app.route('/search_subjects')
def search_subjects_route():
    keyword = request.args.get('keyword', '').strip()
    matched_subjects = search_subjects(keyword)
    return jsonify({'subjects': matched_subjects})

@app.route('/tracnghiem', methods=["POST", "GET"])
def nhapma():
    if "user" in session:
        if request.method == "GET":
            return render_template('tracnghiem.html')
        else:
            monthi=request.form["monhoc"]
            search_result = search_name_in_database(monthi)
            if not search_result:
                flash(f"Chưa có đề trắc nghiệm cho môn này", category="info")
                return redirect(url_for("nhapma"))

            else:
                questions.clear()
                for row in search_result:
                    res={}
                    res['question']=row[3]
                    options=[]
                    options.append(row[4])
                    options.append(row[5])
                    options.append(row[6])
                    options.append(row[7])
                    res['options']=options
                    res['correctAnswer']=row[8]
                    questions.append(res)
                return redirect(url_for("get_questions"))

    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))


#thi trắc nghiệm:
@app.route('/thitracnghiem')
def get_questions():
    if "user" in session:
        return render_template('thitracnghiem.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
#them cau hoi vao phan data tren web
@app.route('/data')
def get_data():
    if "user" in session:
        return jsonify(questions)
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))


# Thêm câu hỏi từ web vào database phần này danh cho admin
@app.route('/themmon')
def themde():
    if "user" in session:
        return render_template("themmon.html")
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/save-data', methods=['POST'])
def save_data():
    if "user" in session:
        data = request.json  # Lấy dữ liệu được gửi từ trang web
        # Đoạn mã xử lý dữ liệu ở đây (ví dụ: lưu vào cơ sở dữ liệu)
        conn=mysql.connector.connect(user='root',password='123456',host='localhost')
        cursor = conn.cursor()
        for i in data.get('questions'):
            question_text = i.get('question_text')
            option1 = i.get('options')[0]
            option2 = i.get('options')[1]
            option3 = i.get('options')[2]
            option4 = i.get('options')[3]
            correct = i.get('correct_answer')
            if correct=='option1':
                correct_answer=option1
            elif correct=='option2':
                correct_answer=option2
            elif correct=='option3':
                correct_answer=option3
            else:
                correct_answer=option4
            # Thực hiện câu lệnh SQL INSERT INTO
            cursor.execute("INSERT INTO `btlweb`.`nganhangcauhoi` (`mamon`, `tenmon`, `question`, `option1`, `option2`, `option3`, `option4`, `correctAnswer`) VALUES (%s, %s, %s, %s, %s, %s, %s,%s)",
                    (data.get('subject_code'), data.get('subject_name'), question_text, option1, option2, option3, option4, correct_answer))
            conn.commit()
        # Sau đó, trả về phản hồi cho trang web
        response = {'message': 'Data received successfully', 'data': data}
        return jsonify(response)
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

# thêm câu hỏi từ file excel vào CSDL
# @app.route('/upload_excel', methods=['POST'])
# def upload_excel():
#     file = request.files['file']
#     if file.filename == '':
#         return 'No selected file'

#     if file:
#         filename = file.filename
#         a=request.form["mamonhoc"]
#         b=request.form["tenmonhoc"]
#         # Lưu dữ liệu file vào cơ sở dữ liệu
#         # save_to_database(filename, data)
#         df = pd.read_excel(file)
#         conn=mysql.connector.connect(user='root',password='123456',host='localhost')
#         cursor = conn.cursor()
#         for index, row in df.iterrows():
#             if row.iloc[5]=="a" or row.iloc[5]=="A":
#                 correct=row.iloc[1]
#             if row.iloc[5]=="b" or row.iloc[5]=="B":
#                 correct=row.iloc[2]
#             if row.iloc[5]=="c" or row.iloc[5]=="C":
#                 correct=row.iloc[3]
#             if row.iloc[5]=="d" or row.iloc[5]=="D":
#                 correct=row.iloc[4]
#             cursor.execute("INSERT INTO `btlweb`.`nganhangcauhoi` (`mamon`, `tenmon`, `question`, `option1`, `option2`, `option3`, `option4`, `correctAnswer`) VALUES (%s, %s, %s, %s, %s, %s, %s,%s)",
#             (a,b,row.iloc[0],row.iloc[1],row.iloc[2],row.iloc[3],row.iloc[4],correct))
#             conn.commit()    
#         flash("File đã được tải lên và lưu trữ trong cơ sở dữ liệu thành công!",category="info")
#         return redirect(url_for("admin"))
@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'

    if file:
        a = request.form["mamonhoc"]
        b = request.form["tenmonhoc"]
        
        # Mở file Excel
        wb = load_workbook(file)
        sheet = wb.active
        # Lặp qua từng hàng trong sheet
        for row in sheet.iter_rows(values_only=True):
            correct = row[5].lower() if row[5] else None
            cursor.execute("INSERT INTO `btlweb`.`nganhangcauhoi` (`mamon`, `tenmon`, `question`, `option1`, `option2`, `option3`, `option4`, `correctAnswer`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                           (a, b, row[0], row[1], row[2], row[3], row[4], correct))
            conn.commit()

        flash("File đã được tải lên và lưu trữ trong cơ sở dữ liệu thành công!", category="info")
        return redirect(url_for("admin"))

#xử lý đăng ký tại đây
def check1(a):
    cursor.execute("SELECT * FROM btlweb.nguoidung")
    result = cursor.fetchall()
    if result is not None:
        for row in result:
            if row[1]==a:
                return 0
    return 1

@app.route('/dangky', methods=["POST", "GET"] )
def dangky():
    if request.method=="POST":
        tendangnhap=request.form["username"]
        matkhau=request.form["password"]
        if check1(tendangnhap)!=0:
            cursor.execute("INSERT INTO `btlweb`.`nguoidung` (`username`, `pass`) VALUES (%s, %s)",(tendangnhap,matkhau))
            conn.commit()
            flash("Bạn đăng ký tài khoản thành công",category="info")
            return redirect(url_for("index"))
        else:
            flash("Tài khoản bạn đăng ký đã tồn tại",category="info")
            return render_template("dangky.html")
    else:
        return render_template("dangky.html")
    
#code cho phan logout
@app.route('/logout')
def log_out():
    session.pop("user", None)
    return redirect(url_for("index"))

# lấy lại mật khẩu

# hàm để gửi email
def send_email(email, verification_code):
    smtp_server = 'smtp.googlemail.com'
    smtp_port = 587
    sender_email = 'sieugazl02@gmail.com'
    sender_password = 'chyr ktgf sbmt uhcg'

    message = MIMEText(f'Mã xác nhận của bạn là: {verification_code}')
    message['Subject'] = 'WEB HỌC TẬP NHÓM 4'
    message['From'] = sender_email
    message['To'] = email

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, [email], message.as_string())
#ham gửi email thông bao về cập nhật câu hỏi:
# def send_email1(email, tenmon , mamon):
#     smtp_server = 'smtp.googlemail.com'
#     smtp_port = 587
#     sender_email = 'sieugazl02@gmail.com'
#     sender_password = 'chyr ktgf sbmt uhcg'

#     message = MIMEText(f'Bộ câu hỏi của môn {tenmon} vừa được cập nhật. Bạn có thể vào phần thực hành nhập mã {mamon} để làm ngay.')
#     message['Subject'] = 'WEB HỌC TẬP NHÓM 4'
#     message['From'] = sender_email
#     message['To'] = email

#     with smtplib.SMTP(smtp_server, smtp_port) as server:
#         server.starttls()
#         server.login(sender_email, sender_password)
#         server.sendmail(sender_email, [email], message.as_string())        
        
# nhập email để lấy lại mật khẩu
@app.route('/quenmatkhau1', methods=["POST","GET"])
def verify_email():
    if request.method=="POST":
        email = request.form['email']
        session['email'] = email
        # tạo mã xác nhận
        verification_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        session['verification_code'] = verification_code

        # gửi mã xác nhận cho email
        send_email(email, verification_code)
        return redirect(url_for("confirm_code"))
    else:
        return render_template('quenmatkhau1.html')

@app.route('/quenmatkhau2', methods=['POST','GET'])
def confirm_code():
    if request.method=="POST":
        entered_code = request.form['code']
        if entered_code == session.get('verification_code'):
            return redirect(url_for("newpassword")) 
        else:
            flash("Mã xác nhận của bạn không đúng",category="info")
            return redirect(url_for("verify_email")) 
    else:
        return  render_template('quenmatkhau2.html')

@app.route('/quenmatkhau3', methods=['POST','GET'])
def newpassword():
    if request.method=="POST":
        mk1 = request.form['password1']
        mk2 = request.form['password2']
        if mk1==mk2:
            #cap nhat mk cho sql
            cursor.execute("SELECT * FROM btlweb.nguoidung where username=%s;",(session['email'],))
            result=cursor.fetchall()
            a=result[0][0]
            cursor.execute("UPDATE btlweb.nguoidung SET pass=%s WHERE userid = %s",(mk1,a))
            conn.commit()
            flash("Lấy lại mật khẩu thành công",category="info")
            return redirect(url_for("index"))
        else:
            flash("Mật khẩu nhập lại không trùng khớp",category="info")
            return redirect(url_for("newpassword"))
    else:
        return render_template('quenmatkhau3.html')



@app.route('/webdesign')
def webdesign():
    if "user" in session:
        return render_template('web_design.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/gioi_thieu_HTML')
def HTML0():
    if "user" in session:
        return render_template('HTML_la_gi.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
    

#luyện code HTML

@app.route('/codehtml')
def codehtml():
    if "user" in session:
        return render_template('codehtml.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
    
@app.route('/HTML_la_gi')
def HTML1():
    if "user" in session:
        return render_template('HTML_la_gi.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/HTML_hoat_dong_the_nao')
def HTML2():
    if "user" in session:
        return render_template('HTML_hoat_dong_the_nao.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/HTML_thuat_ngu_thuong_dung')
def HTML3():
    if "user" in session:
        return render_template('HTML_thuat_ngu_thuong_dung.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/hoc_ngon_ngu_nao')
def HTML4():
    if "user" in session:
        return render_template('/hoc_ngon_ngu_nao.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/gioi_thieu_css1')
def CSS1():
    if "user" in session:
        return render_template('/gioithieuCSS1.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/gioi_thieu_css2')
def CSS2():
    if "user" in session:
        return render_template('/gioithieuCSS2.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/gioi_thieu_css3')
def CSS3():
    if "user" in session:
        return render_template('/gioithieuCSS3.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
        
@app.route('/gioi_thieu_dsa1')
def DSA1():
    if "user" in session:
        return render_template('/gioithieuDSA1.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

@app.route('/gioi_thieu_dsa2')
def DSA2():
    if "user" in session:
        return render_template('/gioithieuDSA2.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
#luyện code js
@app.route('/codejs')
def codejs():
    if "user" in session:
        return render_template('codejs.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))
    
@app.route('/gioi_thieu_JS')
def JS1():
    if "user" in session:
        return render_template('/gioithieuJS.html')
    else:
        flash("Bạn hãy đăng nhập tài khoản",category="info")
        return redirect(url_for("index"))

# phần này dành cho chat bot  

clf = pickle.load(open('Colab\\NB_ChatBot_model.pkl', 'rb'))
vocabulary_to_load = pickle.load(open('Colab\\vocab.pkl', 'rb'))
le = pickle.load(open('Colab\\decode_label.pkl', 'rb'))

@app.route("/chatbot", methods=["POST"])
def chatbot_response():
    if request.method == "POST":
        message = request.form.get("msg")
        ok = prediction(message)
    return ok


# def tienxuly(document):
#     document = ViTokenizer.tokenize(document)
#     # đưa về lower
#     document = document.lower()
#     # xóa các ký tự không cần thiết
#     document = re.sub(r'[^\s\wáàảãạăắằẳẵặâấầẩẫậéèẻẽẹêếềểễệóòỏõọôốồổỗộơớờởỡợíìỉĩịúùủũụưứừửữựýỳỷỹỵđ_]',' ',document)
#     # xóa khoảng trắng thừa
#     document = re.sub(r'\s+', ' ', document).strip()
#     return document
def tienxuly(document):
    # Tách từ bằng biểu thức chính quy
    document = re.findall(r'\b\w+\b', document)
    # Đưa về lower
    document = [word.lower() for word in document]
    # Loại bỏ các ký tự không mong muốn và dấu câu
    document = [re.sub(r'[^\w\s]', '', word) for word in document]
    # Xóa khoảng trắng thừa
    document = ' '.join(document).strip()
    return document


stopword = ["bot","ra"]


def remove_stopwords(line):
    words = []
    for word in line.strip().split():
        if word not in stopword:
            words.append(word)
    return ' '.join(words)


def prediction(input):
    ngram_size = 1
    loaded_vectorizer = CountVectorizer(ngram_range=(ngram_size, ngram_size), min_df=1,
                                        vocabulary=vocabulary_to_load)
    loaded_vectorizer._validate_vocabulary()
    a = tienxuly(input)

    input1 = remove_stopwords(a)
    vect = loaded_vectorizer.transform([input1]).toarray()
    predict = clf.predict(vect)
    predict = le.inverse_transform(predict)[0]
    
    if predict=="noanswer":
        predict= 'Xin lỗi bạn, câu này mình không biết trả lời như thế nào. Bạn vui lòng liện hệ theo số điện thoại 123456 để được tư vấn trực tiếp hoặc nhấn vào nút "Trợ giúp%s" ở góc trái trên cùng màn hình để được giúp đỡ.'

    return predict





if __name__ == "__main__":
    
    app.run(debug=True)
